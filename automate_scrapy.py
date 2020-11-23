"""
自动化爬虫
ASIN	产品图片	品牌名	店铺名	产品上新时间	产品评论数	大类排名	小类排名	序号	店铺名	店铺30天feeback数	店铺总feeback数	店铺产品数	品牌名搜索结果数	店铺品牌
"""
import openpyxl
import requests
import json
import time
import re
import os
import base64
import shutil

from io import BytesIO
from pyquery import PyQuery as pq
from openpyxl.drawing.image import Image as op_img
from selenium import webdriver
from selenium.webdriver import DesiredCapabilities
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By


class Category(object):

    def __init__(self, url_main, save_name):
        self.postal_dict = {
            "https://www.amazon.com/": "20237",
            "https://www.amazon.ca/": "K1V-7P8",
            "https://www.amazon.co.jp/": "163-8001",
            "https://www.amazon.de/": "10115",
            "https://www.amazon.it/": "67061",
            "https://www.amazon.es/": "28080",
            "https://www.amazon.co.uk/": "NW1 6XE",
            "https://www.amazon.fr/": "75020",
        }
        self.urls = url_main  # 主要入口链接
        self.open_file = save_name if os.path.exists(save_name) else "亚马逊派对卖家.xlsx"
        self.wb = openpyxl.load_workbook(self.open_file)
        self.ws = self.wb.active
        self.ws.title = save_name
        self.save_name = save_name  # 导出表格文件名
        # 设置文字图片单元格的行高列宽
        self.column_width = 8
        self.row_height = 50

        options = webdriver.ChromeOptions()  # 初始化浏览器选项
        options.add_argument('--disable-gpu')  # 谷歌文档提到需要加上这个属性来避免bug
        options.add_argument("disable-web-security")  # 允许重定向
        options.add_argument('disable-infobars')  # 隐藏 “Chrome正在受到自动软件的控制”
        # options.add_argument("--headless")  # => 为Chrome配置无头模式
        options.add_experimental_option('excludeSwitches', ['enable-automation'])  # 设置为开发者模式
        # 禁用浏览器弹窗
        prefs = {
            'profile.default_content_setting_values': {
                'images': 2
            }
        }
        options.add_experimental_option('prefs', prefs)
        options.add_argument('blink-settings=imagesEnabled=false')  # 不加载图片，提升速度
        # capa = DesiredCapabilities.CHROME
        # capa["pageLoadStrategy"] = "none"  # 不等待解析完成，直接返回
        self.driver = webdriver.Chrome(options=options)  # 加载浏览器选项
        self.driver.maximize_window()
        self.wait = WebDriverWait(self.driver, 5)

    # 获取ASIN
    def get_asin(self, url: str, start: int, end: int) -> None:
        """
        获取ASIN
        :param url:
        :param start:
        :param end:
        :return:
        """
        options = webdriver.ChromeOptions()
        options.add_argument('--disable-gpu')
        options.add_argument("disable-web-security")
        options.add_argument('disable-infobars')
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        capa = DesiredCapabilities.CHROME
        capa["pageLoadStrategy"] = "none"  # 不等待解析完成，直接返回
        options.add_argument("--headless")  # => 为Chrome配置无头模式
        driver = webdriver.Chrome(options=options, desired_capabilities=capa)
        wait = WebDriverWait(driver, 30)
        driver.maximize_window()
        driver.get(url + f"&page={start}")
        file = open("asin1.txt", "a+", encoding="utf8")
        index = 0
        try:
            time.sleep(3.5)
            driver.find_element_by_css_selector("#sp-cc-accept").click()
        except:
            pass
        for p in range(start, end + 1):
            try:
                print(f"*************正在进行第 {p} 页***************")
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.s-main-slot > div")))
                time.sleep(3)
                div_list = driver.find_elements_by_css_selector("div.s-main-slot > div")
                p_id = 1
                for id, node in enumerate(div_list):
                    asin = node.get_attribute("data-asin")
                    if asin:
                        print(asin)
                        file.write(asin + ",")
                        p_id = id
                index = index + p_id
                time.sleep(2)
                page_num_list = driver.find_element_by_css_selector("ul.a-pagination")
                page_num_list.find_element_by_link_text(str(p)).click()
            except:
                print(f"=========第 {p} 出问题了=============")
                pass
        print("ASIN全部获取成功！")
        file.close()
        driver.quit()

    # 返回ASIN
    def return_asin(self):
        """
        返回ASIN
        :return:
        """
        file = open("asin1.txt", "r", encoding="utf8").read()
        file_list = file.rstrip(",").split(",")
        print('未去重数量:', len(file_list))
        asin_list = {}.fromkeys(file_list).keys()
        return list(asin_list)

    # 验证码识别
    def identify_captcha(self):
        while True:
            try:
                img_url = self.driver.find_element_by_xpath(
                    "/html/body/div/div[1]/div[3]/div/div/form/div[1]/div/div/div[1]/img").get_attribute('src')

                try:
                    img = requests.get(img_url)
                    ls_f = base64.b64encode(BytesIO(img.content).read())
                    # 打印出这个base64编码
                    imgdata = base64.b64decode(ls_f)
                    post_url = 'http://180.76.101.3:7788/'
                    response = requests.post(url=post_url, data=imgdata)
                    text = json.loads(response.text)

                    try:
                        captcha_num = text["code"]
                        print("我在验证码识别这边，现在获取到的验证码是" + str(captcha_num))
                        time.sleep(1)
                        # 输入验证码
                        self.driver.find_element_by_xpath("//*[@id='captchacharacters']").send_keys(captcha_num)
                        # 点击确定按钮
                        self.driver.find_element_by_xpath(
                            "/html/body/div/div[1]/div[3]/div/div/form/div[2]/div/span").click()
                    except:
                        print("---------获取验证码失败---------")
                        return self.identify_captcha()
                except requests.exceptions.ConnectionError as e:
                    print("捕捉到requests.exceptions.ConnectionError 异常 %s" % e, "重新请求")
                    return self.identify_captcha()
            except Exception as e:
                break

    # 更改地址
    def change_address(self, postal, shop_url):
        while True:
            if 'co.jp' in shop_url or '.ca' in shop_url:
                try:
                    self.wait.until(EC.presence_of_element_located((By.ID, 'glow-ingress-line1'))).click()
                    time.sleep(2)
                except Exception as e:
                    self.driver.refresh()
                    continue
                try:
                    time.sleep(1)
                    self.wait.until(EC.presence_of_element_located((By.ID, 'GLUXZipUpdateInput_0'))).send_keys(
                        postal.split('-')[0])
                    time.sleep(1)
                    self.wait.until(EC.presence_of_element_located((By.ID, 'GLUXZipUpdateInput_1'))).send_keys(
                        postal.split('-')[1])
                    time.sleep(1)
                    self.driver.find_element_by_id('GLUXZipUpdate').click()
                    break
                except Exception as NoSuchElementException:
                    self.driver.find_element_by_id('GLUXZipUpdate').click()
                    time.sleep(2)
                    self.driver.refresh()
                    continue
            else:
                try:
                    self.wait.until(EC.presence_of_element_located((By.ID, 'glow-ingress-line1'))).click()
                    time.sleep(2)
                except Exception as e:
                    self.driver.refresh()
                    continue
                try:
                    WebDriverWait(self.driver, 1).until(
                        #  判断某个元素是否可见. 可见代表元素非隐藏
                        EC.visibility_of_element_located((By.CSS_SELECTOR, 'a#GLUXChangePostalCodeLink')))
                    break
                except Exception as e:
                    pass
                try:
                    self.wait.until(EC.presence_of_element_located((By.ID, "GLUXZipUpdateInput"))).send_keys(postal)
                    time.sleep(1)
                    self.driver.find_element_by_id('GLUXZipUpdate').click()
                    break
                except Exception as e:
                    self.driver.refresh()
                    continue
        time.sleep(2)
        self.driver.refresh()
        return True

    # 保存excel
    def save_excel(self, i, data_dict):
        self.ws.cell(row=i + 1, column=2).value = data_dict['asin']  # asin
        self.ws.column_dimensions['C'].width = self.column_width
        self.ws.row_dimensions[i + 1].height = self.row_height
        try:
            f_name = self.save_img(data_dict['img_src'])
            img = op_img(f_name)
            newsize = (50, 50)
            img.width, img.height = newsize  # 这两个属性分别是对应添加图片的宽高
            self.ws.add_image(img, 'C' + str(i + 1))  # 插入图片
        except Exception as e:
            pass
        self.ws.cell(row=i + 1, column=4).value = data_dict['brand_name']  # 品牌名
        self.ws.cell(row=i + 1, column=5).value = data_dict['shop_name']  # 店铺名
        self.ws.cell(row=i + 1, column=6).value = data_dict['review_num']  # 产品评论数
        self.ws.cell(row=i + 1, column=7).value = data_dict['big']  # 大类排名
        self.ws.cell(row=i + 1, column=8).value = data_dict['small']  # 小类排名
        self.ws.cell(row=i + 1, column=9).value = data_dict['thirty_feeback_count']  # 店铺30天feeback数
        self.ws.cell(row=i + 1, column=10).value = data_dict['total_feeback']  # 店铺总feeback数
        self.ws.cell(row=i + 1, column=11).value = data_dict['good_count']  # 店铺产品数
        self.ws.cell(row=i + 1, column=12).value = data_dict['brand_count']  # 品牌名搜索结果数
        self.ws.cell(row=i + 1, column=13).value = data_dict['shop_url']  # 店铺链接
        self.ws.cell(row=i + 1, column=14).value = data_dict['count_90']  # 90天
        self.ws.cell(row=i + 1, column=15).value = data_dict['count_12']  # 1年
        self.ws.cell(row=i + 1, column=16).value = data_dict['shop_desc']  # 店铺信息
        self.ws.cell(row=i + 1, column=17).value = data_dict['brand_list']  # 品牌列表
        self.ws.cell(row=i + 1, column=18).value = data_dict['price']  # 价格
        self.ws.cell(row=i + 1, column=19).value = data_dict['score']  # 评分
        self.wb.save(self.save_name)

    # 解析数据
    def parse_detail(self, data_dict):
        cur_url = self.driver.current_url + '/r'
        title = self.driver.title
        if "Page Not Found" in title or "404" in title or "Sorry" in title or "Documento" in title or "Die Website" in title or title.startswith("Impossibile"):
            return False
        try:
            asin_regex = re.compile("dp/(.*?)/r")
            data_dict['asin'] = asin_regex.findall(cur_url)[0]
        except Exception as e:
            data_dict['asin'] = ''
        # time.sleep(2)
        # html = pq(self.driver.page_source)

        # 评分数
        try:
            reviews_str = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#acrCustomerReviewText"))).text
            reviews = re.findall(r'(\d+)', reviews_str.replace(",", ""))
            data_dict['review_num'] = reviews[0]
        except:
            data_dict['review_num'] = ""

        # 图片列表
        try:
            img = self.driver.find_elements_by_css_selector("#altImages li.item img")[0]
            data_dict['img_src'] = img.get_attribute('src').replace("._AC_US40_.jpg", "._AC_SX100_.jpg")
        except Exception as e:
            data_dict['img_src'] = ""
        # 价格
        try:
            try:
                price_str = self.driver.find_element_by_css_selector("#priceblock_ourprice").text
            except:
                price_str = self.driver.find_element_by_css_selector("#priceblock_saleprice").text
            if "¥" in price_str:
                price = re.findall(r'(\d+)', price_str.replace(",", ""))
            else:
                price = re.findall(r'(\d+.\d+|\d+)', price_str.replace(',', '.'))
            data_dict['price'] = float(price[0])
        except:
            data_dict['price'] = ""

        # 评分
        try:
            score_str = self.driver.find_element_by_css_selector("#acrPopover").get_attribute("title")
            score = re.findall(r'(\d+.\d+|\d+)', score_str.replace(',', '.'))
            data_dict['score'] = float(score[0])
        except:
            data_dict['score'] = ""

        # 品牌名
        try:

            data_dict['brand_name'] = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a#bylineInfo"))).text
        except Exception as e:
            data_dict['brand_name'] = ""

        # 进入品牌详情页
        try:
            brand_url = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a#bylineInfo"))).get_attribute('href')
            js = f"window.open('{brand_url}')"
            self.driver.execute_script(js)
            handles = self.driver.window_handles
            self.driver.switch_to.window(handles[-1])
            self.identify_captcha()  # 识别验证码

            # 品牌名搜索结果数
            try:
                brand_count = self.driver.find_element_by_xpath(
                    '//*[@id="search"]/span/div/span/h1/div/div[1]/div/div/span[1]').text
                brand_count = brand_count.replace(',', '')
                try:
                    brand_count_regex = re.compile(".*?\s(\d+)\s.*?")
                    brand_count = brand_count_regex.findall(brand_count)[0]
                except Exception as e:
                    brand_count_regex = re.compile(".*?(\d+).*?")
                    brand_count = brand_count_regex.findall(brand_count)[0]
            except Exception as e:
                brand_count = 0
            data_dict['brand_count'] = int(brand_count)
            self.driver.close()
            self.driver.switch_to.window(handles[1])
        except Exception as e:
            pass

        # 店铺名
        try:
            shop_name = self.driver.find_elements_by_css_selector("#sellerProfileTriggerId")[0].text
            if shop_name == "":
                shop_name = self.driver.find_elements_by_css_selector("#sellerProfileTriggerId")[1].text
        except Exception as e:
            shop_name = ""
        data_dict['shop_name'] = shop_name

        # 排名
        max_rank = ""
        min_rank = ""
        try:
            # 有表格
            big_small_html = self.driver.find_element_by_css_selector("#productDetails_detailBullets_sections1").text
            if big_small_html:
                s = re.findall(r'#(.*?) ', big_small_html)
                if not s:
                    s = re.findall(r'(\d+|\d+.\d+) (en|in|dans) ', big_small_html)
                max_rank = s[0][0].replace(",", "").replace(".", "").strip()
                min_rank = s[-1][0].replace(",", "").replace(".", "").strip()
        except Exception as e:
            pass
        try:
            if max_rank == "":
                # 无表格
                big_small_html2 = self.driver.find_element_by_css_selector("#detailBullets_feature_div").text
                if big_small_html2:
                    s = re.findall(r'#(.*?) ', big_small_html2)
                    if not s:
                        s = re.findall(r'(\d+|\d+.\d+) (en|in|dans) ', big_small_html2)
                    max_rank = s[0][0].replace(",", "").replace(".", "").strip()
                    min_rank = s[-1][0].replace(",", "").replace(".", "").strip()
        except Exception as e:
            pass
        try:
            data_dict['big'] = int(max_rank)
            data_dict['small'] = int(min_rank)
        except Exception as e:
            data_dict['big'] = 0
            data_dict['small'] = 0

        # 打开店铺详情页
        try:
            seller_pro = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'a#sellerProfileTriggerId'))).get_attribute('href')
            data_dict['shop_url'] = seller_pro
            js = f"window.location='{seller_pro}'"
            self.driver.execute_script(js)
            self.identify_captcha()  # 识别验证码
        except Exception as e:
            data_dict['shop_url'] = ''

        # 店铺30天feeback数
        try:
            thirty_feeback_count = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, '#feedback-summary-table > tbody > tr:nth-child(5) > td:nth-child(2)'))).text
            thirty_feeback_count = thirty_feeback_count.replace(",", "").replace(".", "").strip()
            if thirty_feeback_count == "-" or thirty_feeback_count == '':
                thirty_feeback_count = 0
            data_dict['thirty_feeback_count'] = int(thirty_feeback_count)
        except Exception as e:
            data_dict['thirty_feeback_count'] = 0

        # 90天feedback
        try:
            count_90 = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'tr:nth-of-type(5) td.a-text-right:nth-of-type(3) span'))).text
            count_90 = count_90.replace(",", "").replace(".", "").strip()
            if count_90 == "-" or count_90 == '':
                count_90 = 0
            data_dict['count_90'] = int(count_90)
        except Exception as e:
            data_dict['count_90'] = 0

        # 1年feedback
        try:
            count_12 = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'tr:nth-of-type(5) td:nth-of-type(4) span'))).text
            count_12 = count_12.replace(",", "").replace(".", "").strip()
            if count_12 == "-" or count_12 == '':
                count_12 = 0
            data_dict['count_12'] = int(count_12)
        except Exception as e:
            data_dict['count_12'] = 0

        # 店铺总feeback数
        try:
            total_feeback = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, '#feedback-summary-table > tbody > tr:nth-child(5) > td:nth-child(5)'))).text
            total_feeback = total_feeback.replace(",", "").replace(".", "").strip()
            if total_feeback == "-" or total_feeback == '':
                total_feeback = 0
            data_dict['total_feeback'] = int(total_feeback)
        except Exception as e:
            data_dict['total_feeback'] = 0

        # 店铺信息
        try:
            shop_desc = self.wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, '#seller-profile-container > div.a-row.a-spacing-medium > div > ul'))).text
            shop_desc = shop_desc.strip().replace('\n', '')
        except Exception as e:
            shop_desc = ""
        data_dict['shop_desc'] = shop_desc

        # 点击跳转产品页
        try:
            products_link = self.driver.find_element_by_css_selector('li#products-link').click()
            self.identify_captcha()  # 识别验证码
        except Exception as e:
            pass

        # 店铺产品数
        try:
            good_count = self.driver.find_element_by_xpath(
                '//*[@id="search"]/span/div/span/h1/div/div[1]/div/div/span[1]').text
            good_count = good_count.replace(',', '')
            try:
                good_count_regex = re.compile(".*?\s(\d+)\s.*?")
                good_count = good_count_regex.findall(good_count)[0]
            except Exception as e:
                good_count_regex = re.compile(".*?(\d+).*?")
                good_count = good_count_regex.findall(good_count)[0]
            data_dict['good_count'] = int(good_count)
        except Exception as e:
            data_dict['good_count'] = 0

        # 品牌列表
        try:
            self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '.a-link-normal span.a-size-base.a-color-base')))
            brand_lists = self.driver.find_elements_by_css_selector('.a-link-normal span.a-size-base.a-color-base')
            brand_list = "/".join([ele.text for ele in brand_lists])
        except Exception as e:
            brand_list = ''
        data_dict['brand_list'] = brand_list
        return True

    # 保存图片
    def save_img(self, url):
        requests.packages.urllib3.disable_warnings()
        res = requests.get(url, verify=False)
        print(type(BytesIO(res.content).read()), type(res.content))
        file_name = url[-26:-15] + '.jpg'  # 26, 15
        path = os.path.join('images')
        if not os.path.exists(path):
            os.mkdir(path)
        img_file_path = os.path.join(path, file_name)
        with open(img_file_path, 'wb') as f:
            for data in res.iter_content(128):
                f.write(data)
        return img_file_path

    # 主函数
    def main(self, start=0):
        """
        :param start: 程序开始位置
        :return:
        """
        try:
            data_dict = {}
            # self.get_asin(self.urls, 1, 31)
            self.driver.get(self.urls)
            self.identify_captcha()  # 识别验证码
            site_regex = re.compile("https://www.amazon.*?/")
            site = site_regex.findall(self.urls)[0]  # 站点
            postal = self.postal_dict[site]
            if self.change_address(postal, self.urls):
                print("收货地址更改完成")
            else:
                print("地址打开失败!")

            i = 1
            for index, row in enumerate(self.ws.rows):
                # asin为空 记录下位置
                i = index+1

            print("去重后数量:", len(self.return_asin()))
            for asin in self.return_asin()[start:]:
                print(f'第 {start} 个ASIN: {asin}')
                good_url = site + 'dp/' + asin
                # 打开详情页地址
                js = f"window.open('{good_url}')"
                self.driver.execute_script(js)
                handles = self.driver.window_handles
                self.driver.switch_to.window(handles[-1])

                if self.parse_detail(data_dict):
                    self.save_excel(i, data_dict)
                    i += 1

                # 切换商品列表页
                self.driver.close()
                self.driver.switch_to.window(handles[0])
                start += 1
                if start == 3:
                    break
            shutil.rmtree("images")

        except Exception as e:
            print(e)
            self.main(start)


if __name__ == '__main__':
    url = 'https://www.amazon.com/s?k=balloon&i=party-supplies&rh=n%3A1055398%2Cn%3A723470011&dc&page=400&qid=1602732446&rnid=2941120011&ref=sr_pg_3'
    save_name = 'party-supplies-1.com.xlsx'
    spider = Category(url_main=url, save_name=save_name)
    # spider.get_asin(url, 1, 2)
    spider.main()
    print('抓取完成，即将退出!')
    spider.driver.quit()


