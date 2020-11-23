import base64
import json
import time
from io import BytesIO
import openpyxl
import requests
from lxml import etree
import re
import os
import random
from selenium import webdriver
from selenium.webdriver import DesiredCapabilities
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import shutil
from openpyxl.drawing.image import Image
from PIL import Image as Pg
from typing import Any, Dict, List


class RequestAmazonScrapy:
    
    def __init__(self, url_main, save_name, proxy_data=[]):
        self.urls = url_main  # 主要入口链接
        self.save_name = save_name
        self.open_file = save_name if os.path.exists(save_name) else "亚马逊派对卖家.xlsx"
        self.wb = openpyxl.load_workbook(self.open_file)
        self.ws = self.wb.active
        self.ws.title = save_name
        self.site = re.findall(r'amazon.(.*?)/', url_main)[0]
        self.get_proxy = proxy_data
        self.header = {}

    def get_header(self, url: str, proxy_ip=None) -> None:
        """
        更新请求头 通过验证
        :param url: 请求亚马逊链接
        :param proxy_ip: 代理IP地址
        :return:
        """
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')
        options.add_argument("disable-web-security")
        options.add_argument('disable-infobars')
        if proxy_ip:
            options.add_argument('--proxy-server=http://%s' % proxy_ip)
        caps = {
            'browserName': 'chrome',
            'loggingPrefs': {
                'browser': 'ALL',
                'driver': 'ALL',
                'performance': 'ALL',
            },
            'goog:chromeOptions': {
                'perfLoggingPrefs': {
                    'enableNetwork': True,
                },
                'w3c': False,
            },
        }
        driver = webdriver.Chrome(options=options, desired_capabilities=caps)
        driver.get(url)
        while True:
            try:
                img_url = driver.find_element_by_xpath("/html/body/div/div[1]/div[3]/div/div/form/div[1]/div/div/div[1]/img").get_attribute("src")
                img = requests.get(img_url)
                ls_f = base64.b64encode(BytesIO(img.content).read())
                # 打印出这个base64编码
                img_data = base64.b64decode(ls_f)
                post_url = 'http://180.76.101.3:7788/'
                response = requests.post(url=post_url, data=img_data)
                text = json.loads(response.text)
                captcha_num = text["code"]
                print("我在验证码识别这边，现在获取到的验证码是" + str(captcha_num))
                time.sleep(1)
                # 输入验证码
                driver.find_element_by_xpath("//*[@id='captchacharacters']").send_keys(captcha_num)
                # 点击确定按钮
                driver.find_element_by_xpath(
                    "/html/body/div/div[1]/div[3]/div/div/form/div[2]/div/span").click()
                time.sleep(1)
            except Exception as e:
                break
        cookies = driver.get_cookies()
        cookie = [item["name"] + "=" + item["value"] for item in cookies]
        cookie_str = '; '.join(item for item in cookie)
        header = {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
            "accept-encoding": "gzip, deflate, br",
            "accept-language": "en,zh-CN;q=0.9,zh;q=0.8",
            "cache-control": "no-cache",
            "referer": "https://www.amazon."+self.site+"/",
            "cookie": cookie_str,
            "user-agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.135 Safari/537.36"
        }
        self.header = header
        print("开启自动化,更新请求头～")
        driver.quit()

    def get_asin(self, start: int, end: int) -> None:
        """
        获取ASIN
        :param url:
        :param start:
        :param end:
        :return:
        """
        options = webdriver.ChromeOptions()
        options.add_argument('--disable-gpu')
        options.add_argument("disable-web-security")
        options.add_argument('disable-infobars')
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        capa = DesiredCapabilities.CHROME
        capa["pageLoadStrategy"] = "none"  # 不等待解析完成，直接返回
        options.add_argument("--headless")  # => 为Chrome配置无头模式
        driver = webdriver.Chrome(options=options, desired_capabilities=capa)
        wait = WebDriverWait(driver, 30)
        driver.maximize_window()
        driver.get(self.urls + f"&page={start}")
        file = open("asin1.txt", "a+", encoding="utf8")
        index = 0
        try:
            time.sleep(3)
            driver.find_element_by_css_selector("#sp-cc-accept").click()
        except:
            pass
        for p in range(start, end + 1):
            try:
                print(f"*************正在进行第 {p} 页***************")
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.s-main-slot > div")))
                time.sleep(2)
                div_list = driver.find_elements_by_css_selector("div.s-main-slot > div")
                p_id = 1
                for id, node in enumerate(div_list):
                    asin = node.get_attribute("data-asin")
                    if asin:
                        print(asin)
                        file.write(asin + ",")
                        p_id = id
                index = index + p_id
                time.sleep(2)
                page_num_list = driver.find_element_by_css_selector("ul.a-pagination")
                page_num_list.find_element_by_link_text(str(p)).click()
            except Exception as e:
                print(f"=========第 {p} 出问题了=============")
                pass
        print("ASIN全部获取成功！")
        file.close()
        driver.quit()

    def read_asin(self) -> List:
        """
        返回ASIN
        :return:
        """
        file = open("asin1.txt", "r", encoding="utf8").read()
        file_list = file.rstrip(",").split(",")
        print('未去重数量:', len(file_list))
        asin_list = {}.fromkeys(file_list).keys()
        return list(asin_list)

    def get_data(self, url: str) -> str:
        """
        请求亚马逊拿数据
        :param url: 商品链接
        :return: 返回解析数据
        """
        try:
            if self.get_proxy:
                proxies = random.choice(self.get_proxy)
                proxies_url = {"http://": proxies}
            else:
                proxies = ""
                proxies_url = {}
            response = requests.get(url=url, headers=self.header, proxies=proxies_url, timeout=5).text
            html = etree.HTML(response)
            code = html.xpath('/html/body/div/div[1]/div[3]/div/div/form/div[1]/div/div/div[1]/img/@src')
            if code:
                # 有验证码
                self.get_header(url, proxies)
                response = requests.get(url=url, headers=self.header, proxies=proxies_url, timeout=5).text
                html = etree.HTML(response)
            return html
        except Exception as e:
            print(e)
            self.get_data(url)

    def parse_good_detail_data(self, asin: str) -> List:
        """
        商品详情
        :param asin:
        :return:
        """
        good_url = f"https://www.amazon.{self.site}/dp/{asin}/"
        html = self.get_data(good_url)
        # 可以改
        try:
            title = html.xpath('/html/head/title/text()')[0]
            if "Page Not Found" in title or "404" in title or "Sorry" in title or "Documento" in title or "Die Website" in title or "Impossibile" in title:
                return False
        except:
            return False

        good_asin = asin  # re.findall(r'/dp/(.*?)/', good_url)[0]
        try:
            good_big_img = html.xpath('//*[@id="landingImage"]/@src')[0].strip()
        except:
            good_big_img = ""
        try:
            good_title = html.xpath('//*[@id="productTitle"]/text()')[0].strip()
        except:
            good_title = ""
        try:
            good_star_str = html.xpath('//*[@id="acrPopover"]/@title')[0]
            good_star = re.findall(r'(\d+.\d+|\d+)', good_star_str.replace(',', '.'))[0]
        except:
            good_star = ""
        try:
            good_star_number_str = html.xpath('//*[@id="acrCustomerReviewText"]/text()')[0]
            good_star_number = re.findall(r'(\d+)', good_star_number_str.replace(",", ""))[0]
        except:
            good_star_number = ""
        try:
            good_brand = html.xpath('//*[@id="bylineInfo"]/text()')[0]
        except:
            good_brand = ""
        try:
            try:
                price_str = html.xpath('//*[@id="priceblock_ourprice"]/text()')[0]
            except:
                price_str = html.xpath('//*[@id="priceblock_saleprice"]/text()')[0]
            if "¥" in price_str:
                price = re.findall(r'(\d+)', price_str.replace(",", ""))[0]
            else:
                price = re.findall(r'(\d+.\d+|\d+)', price_str.replace(',', '.'))[0]
        except:
            price = 0
        try:
            # 有表格
            big_small_html = html.xpath('string(//*[@id="productDetails_detailBullets_sections1"])')
            if big_small_html:
                s = re.findall(r'(\d+.*?) (en|in|dans) ', big_small_html)
                if not s:
                    s = re.findall(r'#(.*?) ', big_small_html)
                max_rank = s[0][0].replace(",", "").replace(".", "").strip()
                min_rank = s[-1][0].replace(",", "").replace(".", "").strip()
        except Exception as e:
            pass
        try:
            if max_rank == "":
                # 无表格
                big_small_html2 = html.xpath('string(//*[@id="detailBullets_feature_div"])')
                if big_small_html2:
                    s = re.findall(r'#(.*?) ', big_small_html2)
                    if not s:
                        s = re.findall(r'(\d+.*?) (en|in|dans) ', big_small_html2)
                    max_rank = s[0][0].replace(",", "").replace(".", "").strip()
                    min_rank = s[-1][0].replace(",", "").replace(".", "").strip()
        except Exception as e:
            pass
        try:
            good_big = int(max_rank)
            good_small = int(min_rank)
        except Exception as e:
            good_big = 0
            good_small = 0
        try:
            shop_name = html.xpath('//*[@id="sellerProfileTriggerId"]/text()')[0]
            good_buger_str = html.xpath('//*[@id="sellerProfileTriggerId"]/@href')[0]
            good_buger_id = re.findall(r'seller=(.*?)&', good_buger_str)[0]
            shop_data = self.parse_good_shop_data(good_buger_id)
        except:
            shop_name = ""
            shop_data = ["", "", "", "", "", "", "", "", "", ""]
        return [good_asin, good_big_img, good_brand, shop_name, good_star_number, good_big, good_small, price, good_title, good_star] + shop_data

    def parse_good_shop_data(self, shop_id: str) -> List:
        """
        店铺信息
        :param shop_id:
        :return:
        """
        shop_url = f"https://www.amazon.{self.site}/sp?seller={shop_id}"
        html = self.get_data(shop_url)
        try:
            shop_seller_name = html.xpath('//*[@id="seller-profile-container"]/div[2]/div/ul/li[1]/span/text()')[0]
        except:
            shop_seller_name = ""
        try:
            shop_addr = html.xpath('string(//*[@id="seller-profile-container"]/div[2]/div/ul/li[5]/span/ul)')
        except:
            shop_addr = ""
        try:
            shop_star_str = html.xpath('//*[@id="seller-feedback-summary"]/i[1]/span/text()')[0]
            shop_star = re.findall(r'(\d+.\d+|\d+)', shop_star_str.replace(",", "."))[0]
        except:
            shop_star = ""
        try:
            shop_3_count = html.xpath('//*[@id="feedback-summary-table"]/tr[5]/td[2]/span/text()')[0]
        except:
            shop_3_count = ""
        try:
            shop_9_count = html.xpath('//*[@id="feedback-summary-table"]/tr[5]/td[3]/span/text()')[0].replace(",", "").replace(".", "")
        except:
            shop_9_count = ""
        try:
            shop_12_count = html.xpath('//*[@id="feedback-summary-table"]/tr[5]/td[4]/span/text()')[0].replace(",", "").replace(".", "")
        except:
            shop_12_count = ""
        try:
            shop_all_count = html.xpath('//*[@id="feedback-summary-table"]/tr[5]/td[5]/span/text()')[0].replace(",", "").replace(".", "")
        except:
            shop_all_count = ""
        try:
            me_href = html.xpath('//*[@id="products-link"]/a/@href')[0]
            me_id = re.findall(r'./(.*?)\?', me_href)[0]
            me_data = self.parse_shop_item_data(me_id)  # 请求获取商品数量和品牌名
        except:
            me_data = ["", ""]
        return [shop_seller_name, shop_addr, shop_star, shop_3_count, shop_9_count, shop_12_count, shop_all_count, shop_url]+me_data

    def parse_shop_item_data(self, me_id: str) -> List:
        """
        获取商品列表 商品数量 品牌名
        :param me_id: 店铺ID
        :return: 商品数量 品牌名
        """
        me_url = f"https://www.amazon.{self.site}/s?me={me_id}"
        html = self.get_data(me_url)
        try:
            good_number_text = html.xpath('//*[@id="search"]/span/div/span/h1/div/div[1]/div/div/span[1]/text()')[0]
            good_number = re.findall(r'\s(\d.*?) results', good_number_text.replace(',', '').replace('.', ''))[0]
        except:
            good_number = ""
        try:
            brand_list = html.xpath('//*[@id="brandsRefinements"]/ul//li[@class="a-spacing-micro"]/span/a/span/text()')
            brand_list = "/".join(brand_list)
        except:
            brand_list = ""
        return [good_number, brand_list]

    def save_excel(self, i: int, data_list: List) -> None:
        """
        保存成Excel 由于openpyxl 必须从本地读取图片
        :param i: 写入的行
        :param data_list:写入数据
        :return:
        """
        self.ws.cell(row=i + 1, column=2).value = data_list[0]  # asin
        self.ws.column_dimensions['C'].width = 8
        self.ws.row_dimensions[i + 1].height = 50
        try:
            f_name = self.save_img(data_list[1])
            img = Image(f_name)
            newsize = (50, 50)
            img.width, img.height = newsize  # 这两个属性分别是对应添加图片的宽高
            self.ws.add_image(img, 'C' + str(i + 1))  # 插入图片
        except Exception as e:
            pass
        self.ws.cell(row=i + 1, column=4).value = data_list[2]    # 品牌名
        self.ws.cell(row=i + 1, column=5).value = data_list[3]    # 店铺名
        self.ws.cell(row=i + 1, column=6).value = data_list[4]    # 产品评论数
        self.ws.cell(row=i + 1, column=7).value = data_list[5]    # 大类排名
        self.ws.cell(row=i + 1, column=8).value = data_list[6]    # 小类排名
        self.ws.cell(row=i + 1, column=9).value = data_list[13]   # 店铺30天feeback数
        self.ws.cell(row=i + 1, column=10).value = data_list[16]  # 店铺总feeback数
        self.ws.cell(row=i + 1, column=11).value = data_list[18]  # 店铺产品数
        self.ws.cell(row=i + 1, column=12).value = ""             # 品牌名搜索结果数
        self.ws.cell(row=i + 1, column=13).value = data_list[17]  # 店铺链接
        self.ws.cell(row=i + 1, column=14).value = data_list[14]  # 90天
        self.ws.cell(row=i + 1, column=15).value = data_list[15]  # 1年
        self.ws.cell(row=i + 1, column=16).value = data_list[11]  # 店铺信息
        self.ws.cell(row=i + 1, column=17).value = data_list[19]  # 品牌列表
        self.ws.cell(row=i + 1, column=18).value = data_list[7]   # 价格
        self.ws.cell(row=i + 1, column=19).value = data_list[9]   # 评分
        self.wb.save(self.save_name)
    
    def save_img(self, url: str) -> str:
        """
        保存图片
        :param url:
        :return:
        """
        if url.startswith("http"):
            requests.packages.urllib3.disable_warnings()
            res = requests.get(url, verify=False)
            file_name = url[-26:-15] + '.jpg'  # 26, 15
            path = os.path.join('images')
            if not os.path.exists(path):
                os.mkdir(path)
            img_file_path = os.path.join(path, file_name)
            with open(img_file_path, 'wb') as f:
                for data in res.iter_content(128):
                    f.write(data)
        else:
            b64_head, b64_data = url.split(';base64,')
            img_type = b64_head.split("/")[1]
            data = base64.b64decode(b64_data)
            path = os.path.join('images')
            if not os.path.exists(path):
                os.mkdir(path)
            img_file_path = os.path.join(path,  str(time.time())+"."+img_type)
            with open(img_file_path, 'wb') as f:
                f.write(data)
            if img_type == "webp":
                im = Pg.open(img_file_path)
                if im.mode == "RGBA":
                    im.load()
                    background = Pg.new("RGB", im.size, (255, 255, 255))
                    background.paste(im, mask=im.split()[3])
                    im = background
                im.save(img_file_path, "JPEG")
        return img_file_path
    
    def main(self, start=0, end=None) -> None:
        max_length = self.ws.max_row
        i = max_length
        for asin in self.read_asin()[start:end]:
            data = self.parse_good_detail_data(asin)
            if data:
                self.save_excel(i, data)
                i += 1
            start += 1
            print(f"{start} {asin}")
        shutil.rmtree("images")


def parse_proxy_file(file: str) -> List:
    """
    读取代理文件
    """
    with open(file, 'r') as f:
        proxy = f.read().strip().split('\n')
    return proxy


if __name__ == '__main__':
    # ip_list = parse_proxy_file("ip.txt")
    s = RequestAmazonScrapy("https://www.amazon.fr/s?k=Ballon&rh=n%3A363713031%2Cn%3A14507453031&dc&__mk_fr_FR=%C3%85M%C3%85%C5%BD%C3%95%C3%91&qid=1603850611&rnid=1703605031&ref=sr_nr_n_4", "法国1.xlsx")
    # s.get_asin(1, 2)
    s.main()
